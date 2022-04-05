import logging
import time
import xlrd
import os
import io
import json
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
import xlutils as xlu

try:
    from odoo.tools.misc import xlsxwriter
except ImportError:
    import xlsxwriter
from odoo.tools import date_utils
from odoo import fields, models, api
from odoo.exceptions import ValidationError
from odoo.tools.misc import xlsxwriter
from odoo.http import content_disposition, request
from dateutil.relativedelta import relativedelta

_logger = logging.getLogger(__name__)


class FinancialStatementsReport(models.TransientModel):
    _name = 'financial.statement.financial.statements.report.wizard'
    _description = 'Etats financiers'

    name = fields.Char(string="Nom")
    start_date = fields.Date(string="Date Debut", required=True, default=lambda self: self._init_date_from())
    end_date = fields.Date(string="Date Fin", required=True, default=lambda self: self._init_date_to())

    def _init_date_from(self):
        today = fields.Date.context_today(self)
        fiscal_date = self.env.user.company_id.compute_fiscalyear_dates(today)
        return fiscal_date['date_from']

    def _init_date_to(self):
        today = fields.Date.context_today(self)
        fiscal_date = self.env.user.company_id.compute_fiscalyear_dates(today)
        return fiscal_date['date_to']

    @api.onchange('date_from', 'date_to')
    def onchange_date(self):
        if self.start_date and self.end_date:
            if self.start_date >= self.end_date:
                raise ValidationError(_('La date de debut doit être inférieur à la date de fin'))

    @api.multi
    def button_export_xlsx(self):
        response = request.make_response(
            None,
            headers=[
                ('Content-Type', 'application/vnd.ms-excel'),
                ('Content-Disposition', content_disposition('test.xlsx'))
            ]
        )
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        workbook.close()
        output.seek(0)
        response.stream.write(output.read())
        output.close()

        response.set_cookie('fileToken', token)
        return response

    def get_template_sheet(self):
        wb = xlsxwriter.Workbook(
            "/mnt/extra-addons/financial_statement_reporting/static/template_excel/etats_financiers.xlsx")
        return wb.worksheets()

    def print_xlsx(self):
        if self.start_date > self.end_date:
            raise ValidationError('Start Date must be less than End Date')

        data = {
            'start_date': self.start_date,
            'end_date': self.end_date,
        }
        return {
            'type': 'ir_actions_financial_statement_report_download',
            'data': {'model': 'financial.statement.financial.statements.report.wizard',
                     'options': json.dumps(data, default=date_utils.json_default),
                     'output_format': 'xlsx',
                     'report_name': 'ETATS FINANCIERS',
                     }
        }

    def get_xlsx_report(self, data, response):
        file = "/mnt/extra-addons/financial_statement_reporting/static/template_excel/template_etats_financiers.xlsx"
        latest_date_end = datetime.strptime(data.get('end_date'), '%Y-%m-%d').date() - relativedelta(years=1)
        objet_lists = [
            {'model': 'financial.statement.income.statement', 'name': 'COMPTE DE RESULTAT', 'range_start': 5,
             'additional_column': 4, 'sheet_index': 6, 'starting_row': 2},
            {'model': 'financial.statement.active.balance.sheet', 'name': 'BILAN PAYSAGE', 'range_start': 4,
             'additional_column': 3, 'sheet_index': 5, 'starting_row': 3},
            {'model': 'financial.statement.passive.balance.sheet', 'name': 'BILAN PAYSAGE', 'range_start': 11,
             'additional_column': 10, 'sheet_index': 5, 'starting_row': 3, 'blank_line': 25},
        ]
        DEFAULT_WORD = {'COMPANY_NAME': self.env.user.company_id.name,
                        'COMPANY_PHONE': self.env.user.company_id.phone,
                        'COMPANY_DEPOT_CENTER': self.env.user.company_id.name,
                        'COMPANY_SIGLE': self.env.user.company_id.name,
                        'DATE_START': datetime.strptime(data.get('start_date'), '%Y-%m-%d').strftime('%d/%m/%Y'),
                        'DATE_END': datetime.strptime(data.get('end_date'), '%Y-%m-%d').strftime('%d/%m/%Y'),
                        'COMPANY_REGISTRY': self.env.user.company_id.company_registry,
                        'COMPANY_VAT': self.env.user.company_id.vat,
                        'COMPANY_STREET': self.env.user.company_id.street,
                        'COMPANY_STREET2': self.env.user.company_id.street,
                        'COMPANY_CITY': self.env.user.company_id.city,
                        'COMPANY_COUNTRY': self.env.user.company_id.country_id.name,
                        'NUMBER_MONTH': '12',
                        'LATEST_DATE_END': latest_date_end.strftime("%d/%m/%Y"),
                        }

        with open(file, "rb") as f:
            output = io.BytesIO(f.read())

            workbook = load_workbook(output)
            self._specify_default_value(workbook, DEFAULT_WORD)
            for report in objet_lists:
                sheet = workbook.worksheets[report.get('sheet_index')]
                options = {'date': {'date_from': data.get('start_date'),
                                    'date_to': data.get('end_date'), 'filter': 'this_year'}}
                report_obj = request.env[report.get('model')]

                lines = report_obj._get_item_lines(options)
                for index in range(len(lines)):
                    line = lines[index]
                    columns = line.get('columns')
                    index_column = 1
                    row = index + report.get('starting_row') + 1 if report.get('blank_line') and report.get(
                        'blank_line') < (index + report.get('starting_row')) else index + report.get('starting_row')

                    for col in range(report.get('range_start'), len(columns) + report.get('additional_column')):
                        sheet.cell(row=row, column=col).value = columns[index_column].get('name')
                        index_column += 1

            workbook.save(output)
            workbook.close()
            output.seek(0)
            response.stream.write(output.getvalue())
            output.close()

    def _specify_default_value(self, workbook, default_parameters):
        for sheet in workbook.worksheets:
            i = 0
            for r in range(1, sheet.max_row + 1):
                for c in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(r, c).value
                    for key in default_parameters:
                        if cell_value != None and key in str(cell_value):
                            sheet.cell(r, c).value = str(cell_value).replace(key,
                                                                             default_parameters.get(key))
                            i += 1
